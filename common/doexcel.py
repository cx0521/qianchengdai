#-*-conding:utf-8-*-
#@Time:2019/1/18
#@Author:xiaochen
#@File:doexcel.py
import openpyxl

class DoExcel:
    """这是一个读取/写回excel表中数据的类"""
    def __init__(self,file_name): #初始化里面传sheet_name，意思是只能操作当前表
        #异常处理
        try:
            self.file_name=file_name
            #使用openpyxl.load_workbook函数，对excel表进行读取，再实例化为workbook对象
            self.workbook=openpyxl.load_workbook(filename=file_name)
        #文件未找到异常处理
        except FileNotFoundError as e:
            print("{}not found , please check file_path".format(file_name))
            raise e
    def read(self,sheet_name):
        """这是读取excel表中数据的函数"""
        sheet_name=sheet_name
        #根据表名来获取sheet_name名
        sheet=self.workbook[sheet_name]
        #获取当前sheet_name下的最大行
        max_row=sheet.max_row
        #利用for循环，从第二行开始，把excel表中的测试用例读取出来
        data=[]
        for i in range(2,max_row+1):
            row_data={}
            row_data["case_id"]=sheet.cell(i,1).value
            row_data["title"] = sheet.cell(i, 2).value
            row_data["url"] = sheet.cell(i, 3).value
            row_data["data"] = sheet.cell(i, 4).value
            row_data["method"] = sheet.cell(i, 5).value
            row_data["expectedresult"] = sheet.cell(i, 6).value
            data.append(row_data)
        return data

    def write_back(self,sheet_name,row,actualresult,testresult): #因为初始化函数里面只有文件名
        """这是写回到excel表中测试数据函数"""
        #根据表名来获取sheet_name名
        sheet=self.workbook[sheet_name]
        sheet.cell(row,7).value=actualresult #写回实际结果
        sheet.cell(row,8).value=testresult #写回测试结果
        self.workbook.save(filename=self.file_name) #写完，记得保存


if __name__=="__main__":
    from common.contants import cases_dir
    cases=DoExcel(cases_dir).read("login")
    print(cases)










